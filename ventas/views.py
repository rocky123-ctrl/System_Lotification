from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from decimal import Decimal
import math

from .models import Venta, LiquidacionComision, Cotizacion
from .serializers import VentaSerializer, LiquidacionComisionSerializer, CotizacionSerializer

from django.db.models import Sum, Q
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.pagination import PageNumberPagination
from dateutil.relativedelta import relativedelta

class VentasPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class VentaViewSet(viewsets.ModelViewSet):
    """
    ModelViewSet para procesar ventas de Lotes hechas por Clientes.
    Permite operaciones CRUD y expone un endpoint de cálculo avanzado.
    """
    queryset = Venta.objects.all()
    serializer_class = VentaSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = VentasPagination

    def get_queryset(self):
        user = self.request.user
        all_sales = self.request.query_params.get('all') == 'true'
        
        # Un Superadmin o Administrador puede ver todas las ventas si usa el parámetro 'all'
        is_super = user.is_superuser or (
            hasattr(user, 'user_roles') and 
            user.user_roles.filter(role__name__in=['Superadmin', 'Administrador'], is_active=True).exists()
        )

        if is_super and all_sales:
            queryset = Venta.objects.all().select_related('cliente', 'lote', 'vendedor')
        else:
            # Usuarios normales solo ven sus ventas
            queryset = Venta.objects.filter(vendedor=user).select_related('cliente', 'lote', 'vendedor')

        # Filtros
        anio = self.request.query_params.get('anio')
        mes = self.request.query_params.get('mes')
        search = self.request.query_params.get('search')
        lotificacion = self.request.query_params.get('lotificacion')
        estado = self.request.query_params.get('estado')

        if anio:
            queryset = queryset.filter(fecha_creacion__year=anio)
        if mes:
            queryset = queryset.filter(fecha_creacion__month=mes)
        if lotificacion:
            queryset = queryset.filter(lote__manzana__lotificacion_id=lotificacion)
        if estado:
            if estado == 'ACTIVAS':
                queryset = queryset.filter(estado__in=['GENERADA', 'COMPLETADA'])
            else:
                queryset = queryset.filter(estado=estado)
        if search:
            queryset = queryset.filter(
                Q(cliente__nombres__icontains=search) |
                Q(cliente__apellidos__icontains=search) |
                Q(lote__numero_lote__icontains=search)
            )

        return queryset.order_by('-fecha_creacion')

    def perform_create(self, serializer):
        # Asignar automáticamente el estado según el tipo de pago
        tipo_pago = self.request.data.get('tipo_pago', '').upper()
        estado = 'COMPLETADA' if tipo_pago == 'CONTADO' else 'GENERADA'
        
        # Asignar automáticamente el vendedor como el usuario activo
        serializer.save(vendedor=self.request.user, estado=estado)

    def perform_destroy(self, instance):
        """
        Elimina físicamente el registro de la venta (Hard Delete) en cascada.
        Libera el lote, borra cuotas, servicios y la propia venta de forma permanente.
        """
        from django.db import transaction
        from lotes.models import HistorialLote
        from financiamiento.models import Financiamiento
        from cuentas_cobrar.models import Cuota
        from servicios.models import ConfiguracionServicioLote, PagoServicio
        from ventas.models import Cotizacion
        
        with transaction.atomic():
            lote = instance.lote
            estado_anterior = lote.estado_disponibilidad
            
            # 1. Liberar el lote
            lote.estado_disponibilidad = 'disponible'
            lote.save()
            
            # 2. Registrar en historial de lote
            HistorialLote.objects.create(
                lote=lote,
                estado_disponibilidad_anterior=estado_anterior,
                estado_disponibilidad_nuevo='disponible',
                notas=f"Venta ID {instance.id} ELIMINADA físicamente por {self.request.user.username}. Lote liberado y registros financieros borrados en cascada."
            )
            
            # 3. Eliminar financiamiento y cuotas asociadas al lote
            Financiamiento.objects.filter(lote=lote).delete()
            Cuota.objects.filter(venta=instance).delete()
            
            # 4. Eliminar servicios contratados de ese cliente en ese lote
            ConfiguracionServicioLote.objects.filter(lote=lote).delete()
            PagoServicio.objects.filter(lote=lote).delete()

            # 5. Eliminar cotización asociada a esta venta si existe (se asume que si el mismo cliente cotizó este lote)
            Cotizacion.objects.filter(lote=lote, cliente=instance.cliente).delete()
            
            # 6. Eliminar la venta físicamente
            instance.delete()

    @action(detail=True, methods=['post'])
    def escriturar(self, request, pk=None):
        from django.db import transaction
        from lotes.models import HistorialLote

        instance = self.get_object()
        if instance.estado != 'COMPLETADA':
            return Response({"error": "Solo se pueden escriturar ventas completadas."}, status=status.HTTP_400_BAD_REQUEST)
            
        lote = instance.lote
        if lote.estado_disponibilidad == 'escriturado':
            return Response({"error": "Este lote ya se encuentra escriturado."}, status=status.HTTP_400_BAD_REQUEST)
            
        with transaction.atomic():
            estado_anterior = lote.estado_disponibilidad
            lote.estado_disponibilidad = 'escriturado'
            lote.save()
            
            HistorialLote.objects.create(
                lote=lote,
                estado_disponibilidad_anterior=estado_anterior,
                estado_disponibilidad_nuevo='escriturado',
                notas=f"Lote escriturado desde la venta {instance.id} por {self.request.user.username}."
            )
            
        return Response({"mensaje": "Lote escriturado correctamente.", "nuevo_estado_lote": "escriturado"})

    @action(detail=True, methods=['post'])
    def eliminar_permanente(self, request, pk=None):
        """
        Elimina físicamente el registro de la base de datos.
        Solo permitido si ya está en estado CANCELADA.
        """
        instance = self.get_object()
        if instance.estado != 'CANCELADA':
            return Response({"error": "Solo se pueden eliminar permanentemente las ventas que ya están CANCELADAS."}, status=status.HTTP_400_BAD_REQUEST)
        
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=['get'])
    def resumen(self, request):
        """
        Devuelve totales de ventas y comisiones para el usuario activo
        basado en los filtros aplicados.
        Excluye las ventas CANCELADAS de los totales.
        """
        queryset = self.filter_queryset(self.get_queryset())
        
        # Filtramos explícitamente para excluir canceladas de los cálculos monetarios
        queryset_para_totales = queryset.exclude(estado='CANCELADA')
        
        totales = queryset_para_totales.aggregate(
            total_ventas=Sum('valor_lote'),
            total_comisiones=Sum('comision_monto')
        )
        return Response({
            'total_ventas': totales['total_ventas'] or 0,
            'total_comisiones': totales['total_comisiones'] or 0,
            'conteo': queryset_para_totales.count()
        })

    @action(detail=False, methods=['post'])
    def calcular(self, request):
        """
        Calcula amortización y desglose financiero de una venta
        en tiempo real desde el Backend.
        """
        try:
            valor_lote = Decimal(str(request.data.get('valor_lote', 0)))
            acepta_instalacion = request.data.get('acepta_instalacion', False)
            costo_instalacion = Decimal(str(request.data.get('costo_instalacion', 0)))
            enganche = Decimal(str(request.data.get('enganche', 0)))
            descuento = Decimal(str(request.data.get('descuento', 0)))
            tipo_pago = str(request.data.get('tipo_pago', 'contado')).lower()
            plazo_meses = int(request.data.get('plazo_meses', 0))
            tasa_interes = float(request.data.get('tasa_interes', 0))

            # Extraer enganche neto descontando la instalacion
            enganche_puro = max(Decimal('0'), enganche - costo_instalacion)

            # 1) Valor Lote Neto y con descuento
            valor_lote_con_descuento = max(Decimal('0'), valor_lote - descuento)
            
            # 2) Valor Total = (Valor del lote + el Costo de instalacion)
            # Tomamos en cuenta el descuento sobre el lote primero.
            valor_total = valor_lote_con_descuento + costo_instalacion
            
            # 3) Valor a Financiar = Valor Total - enganche
            valor_financiar = max(Decimal('0'), valor_total - enganche)
            faltante = max(Decimal('0'), valor_total - enganche)

            # 4) Tasa Mensual Efectiva r = (tasa_anual * 0.01) / 12
            tasa_anual_decimal = tasa_interes * 0.01
            tasa_mensual_efectiva = round(tasa_anual_decimal / 12.0, 12)
            tasa_mensual_efectiva_porcentaje = tasa_mensual_efectiva * 100

            # 5) Cuota Final Mensual = (P * r * ((1+ r)^n)) / (((1+ r)^n) - 1)
            cuota_final = 0.0
            if tipo_pago == 'financiado' and plazo_meses > 0 and float(valor_financiar) > 0:
                P = float(valor_financiar)
                r = float(tasa_mensual_efectiva)
                n = float(plazo_meses)
                
                if r > 0:
                    numerador = P * r * math.pow(1 + r, n)
                    denominador = math.pow(1 + r, n) - 1
                    cuota_final = round(numerador / denominador, 2)
                else:
                    cuota_final = round(P / n, 2)

            total_pagar_hoy = float(enganche + valor_financiar) if tipo_pago == 'contado' else float(enganche)

            # 6) Total de Financia (+ intereses) = Cuota Final Mensual * Plazo en meses
            total_financia_mas_intereses = round(cuota_final * plazo_meses, 2)
            
            # 7) Valor de los intereses = Total de Financia - Valor a Financiar
            valor_intereses = round(total_financia_mas_intereses - float(valor_financiar), 2)
            
            enganche_puro = max(Decimal('0'), enganche - costo_instalacion)

            return Response({
                'valor_lote': float(valor_lote),
                'valor_lote_con_descuento': float(valor_lote_con_descuento),
                'valor_total': float(valor_total),
                'valor_con_descuento': float(valor_total),
                'enganche_puro': float(enganche_puro),
                'faltante': float(faltante),
                'valor_financiar': float(valor_financiar),
                'tasa_anual': tasa_interes,
                'tasa_mensual_efectiva_porcentaje': tasa_mensual_efectiva_porcentaje,
                'plazo_meses': plazo_meses,
                'cuota_final_mensual': cuota_final,
                'total_financia_mas_intereses': total_financia_mas_intereses,
                'valor_intereses': valor_intereses,
                'total_pagar_hoy': total_pagar_hoy
            }, status=status.HTTP_200_OK)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def reporte_dashboard(self, request):
        from lotes.models import Lote, Manzana
        from financiamiento.models import Financiamiento, Cuota
        from django.db.models import Sum, Count, Q
        from django.db.models.functions import TruncMonth
        import calendar
        from datetime import date

        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        lotificacion_id = request.query_params.get('lotificacion_id')

        # 1. Filtros base (Históricos / Generales)
        ventas_qs = Venta.objects.exclude(estado='CANCELADA')
        lotes_qs = Lote.objects.all()
        cuotas_qs = Cuota.objects.filter(estado='pagada')

        if lotificacion_id and lotificacion_id != 'all':
            ventas_qs = ventas_qs.filter(lote__manzana__lotificacion_id=lotificacion_id)
            lotes_qs = lotes_qs.filter(manzana__lotificacion_id=lotificacion_id)
            cuotas_qs = cuotas_qs.filter(financiamiento__lote__manzana__lotificacion_id=lotificacion_id)

        # 2. Métricas de Lotes (Estado)
        estados_lotes = lotes_qs.values('estado_disponibilidad').annotate(total=Count('id'))
        dict_estados = {item['estado_disponibilidad']: item['total'] for item in estados_lotes}
        
        lotes_disponibles = dict_estados.get('disponible', 0)
        lotes_financiados = dict_estados.get('financiado', 0)
        lotes_pagados = dict_estados.get('pagado', 0) + dict_estados.get('escriturado', 0)
        lotes_reservados = dict_estados.get('reservado', 0)
        
        # 3. Métricas Financieras Históricas
        # Valor Total de Ventas y Enganches
        ventas_totales = ventas_qs.aggregate(
            total_ventas=Sum('valor_lote'),
            total_enganches=Sum('enganche')
        )
        valor_total_ventas = ventas_totales['total_ventas'] or Decimal('0.00')
        valor_enganches = ventas_totales['total_enganches'] or Decimal('0.00')

        # Nuevo KPI: Valor Total del Proyecto (Suma de los valores de todos los lotes)
        valor_total_proyecto_lotes = lotes_qs.aggregate(total=Sum('valor_total'))['total'] or Decimal('0.00')

        # Capital e Intereses Cobrados
        cobros_totales = cuotas_qs.aggregate(
            capital=Sum('monto_capital'),
            interes=Sum('monto_interes')
        )
        valor_capital_cobrado = cobros_totales['capital'] or Decimal('0.00')
        valor_intereses_cobrados = cobros_totales['interes'] or Decimal('0.00')

        finan_qs = Financiamiento.objects.filter(estado='activo')
        if lotificacion_id and lotificacion_id != 'all':
            finan_qs = finan_qs.filter(lote__manzana__lotificacion_id=lotificacion_id)
        
        saldo_pendiente = finan_qs.aggregate(total=Sum('saldo'))['total'] or Decimal('0.00')

        ventas_contado = ventas_qs.filter(tipo_pago='CONTADO').aggregate(total=Sum('valor_lote'))['total'] or Decimal('0.00')
        valor_reservas = Decimal('0.00')

        # 4. Tendencia Ventas Mensuales (para el gráfico - Afectado por fechas)
        ventas_mensuales_qs = ventas_qs
        if fecha_inicio:
            try:
                y, m = map(int, fecha_inicio.split('-'))
                dt_inicio = date(y, m, 1)
                ventas_mensuales_qs = ventas_mensuales_qs.filter(fecha_creacion__date__gte=dt_inicio)
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                y, m = map(int, fecha_fin.split('-'))
                _, last_day = calendar.monthrange(y, m)
                dt_fin = date(y, m, last_day)
                ventas_mensuales_qs = ventas_mensuales_qs.filter(fecha_creacion__date__lte=dt_fin)
            except ValueError:
                pass

        ventas_mensuales = ventas_mensuales_qs.annotate(mes=TruncMonth('fecha_creacion')).values('mes').annotate(
            ventas=Count('id'),
            monto=Sum('valor_lote')
        ).order_by('mes')

        meses_nombres = {1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun', 
                         7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'}
        
        data_ventas_mensuales = []
        for vm in ventas_mensuales:
            if vm['mes']:
                data_ventas_mensuales.append({
                    'mes': f"{meses_nombres[vm['mes'].month]} {vm['mes'].year}",
                    'ventas': vm['ventas'],
                    'monto': float(vm['monto'] or 0)
                })

        # 5. Resumen por Manzana
        manzanas_qs = Manzana.objects.all()
        if lotificacion_id and lotificacion_id != 'all':
            manzanas_qs = manzanas_qs.filter(lotificacion_id=lotificacion_id)
            
        resumen_por_manzana = []
        for manzana in manzanas_qs:
            l_manzana = lote_qs_manzana = Lote.objects.filter(manzana=manzana)
            v_manzana = ventas_qs.filter(lote__manzana=manzana)
            
            est = l_manzana.values('estado_disponibilidad').annotate(c=Count('id'))
            d_est = {x['estado_disponibilidad']: x['c'] for x in est}
            
            v_tot = v_manzana.aggregate(t=Sum('valor_lote'))['t'] or Decimal('0.00')
            
            resumen_por_manzana.append({
                'manzana': manzana.nombre,
                'disponibles': d_est.get('disponible', 0),
                'financiados': d_est.get('financiado', 0),
                'reservados': d_est.get('reservado', 0),
                'pagados': d_est.get('pagado', 0) + d_est.get('escriturado', 0),
                'valorTotal': float(v_tot)
            })

        # 6. Ventas por vendedor en el periodo
        ventas_por_vendedor_qs = ventas_mensuales_qs.values('vendedor__username', 'vendedor__first_name', 'vendedor__last_name', 'vendedor__empleado__nombre', 'vendedor__empleado__apellido').annotate(
            ventas=Count('id'),
            monto=Sum('valor_lote')
        ).order_by('-ventas')

        ventas_por_vendedor = []
        for v in ventas_por_vendedor_qs:
            if v.get('vendedor__empleado__nombre'):
                nombre = f"{v['vendedor__empleado__nombre']} {v.get('vendedor__empleado__apellido', '')}".strip()
            elif v.get('vendedor__first_name'):
                nombre = f"{v['vendedor__first_name']} {v.get('vendedor__last_name', '')}".strip()
            else:
                nombre = v['vendedor__username']
            
            ventas_por_vendedor.append({
                'vendedor': nombre,
                'ventas': v['ventas'],
                'monto': float(v['monto'] or 0)
            })

        response_data = {
            'lotesDisponibles': lotes_disponibles,
            'lotesFinanciados': lotes_financiados,
            'lotesReservados': lotes_reservados,
            'lotesPagados': lotes_pagados,
            'lotesCancelados': 0, # En lotes el estado vuelve a disponible
            'valorTotalVentas': float(valor_total_ventas),
            'valorTotalProyectoLotes': float(valor_total_proyecto_lotes),
            'valorEnganches': float(valor_enganches),
            'valorCapitalCobrado': float(valor_capital_cobrado + ventas_contado), # Se incluye venta al contado como capital
            'valorInteresesCobrados': float(valor_intereses_cobrados),
            'valorReservas': float(valor_reservas),
            'valorPendienteCobro': float(saldo_pendiente),
            'dataVentasMensuales': data_ventas_mensuales,
            'resumenPorManzana': resumen_por_manzana,
            'ventasPorVendedor': ventas_por_vendedor
        }

        return Response(response_data)

    @action(detail=False, methods=['get'])
    def reporte_financiamiento_clientes(self, request):
        from cuentas_cobrar.models import Cuota
        from django.db.models import Sum, Q

        lotificacion_id = request.query_params.get('lotificacion_id')
        
        ventas_qs = Venta.objects.filter(tipo_pago='FINANCIADO').exclude(estado='CANCELADA')
        if lotificacion_id and lotificacion_id != 'all':
            ventas_qs = ventas_qs.filter(lote__manzana__lotificacion_id=lotificacion_id)
            
        ventas_qs = ventas_qs.select_related('cliente', 'lote').order_by('-fecha_creacion')
        
        cuotas_qs = Cuota.objects.filter(venta__in=ventas_qs)
        totales = cuotas_qs.aggregate(
            total_monto=Sum('monto_cuota'),
            pagados=Sum('monto_cuota', filter=Q(estado='Pagado')),
            pendientes=Sum('monto_cuota', filter=Q(estado__in=['Pendiente', 'Vencido']))
        )
        total_monto = totales['total_monto'] or Decimal('0.00')
        pagados = totales['pagados'] or Decimal('0.00')
        pendientes = totales['pendientes'] or Decimal('0.00')

        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(ventas_qs, request)
        if page is not None:
            data = []
            for venta in page:
                cuotas = venta.cuotas_cobrar.all()
                total_cuotas = cuotas.count()
                pagadas = cuotas.filter(estado='Pagado').count()
                
                data.append({
                    'venta_id': venta.id,
                    'cliente_nombre': f"{venta.cliente.nombres} {venta.cliente.apellidos}".strip(),
                    'lote': venta.lote.numero_lote,
                    'progreso_cuotas': f"{pagadas}/{total_cuotas}"
                })
            
            response = paginator.get_paginated_response(data)
            response.data['totales'] = {
                'total_monto': float(total_monto),
                'pagados': float(pagados),
                'pendientes': float(pendientes)
            }
            return response

        return Response([])

    @action(detail=False, methods=['get'])
    def reporte_servicios_clientes(self, request):
        from servicios.models import ConfiguracionServicioLote, PagoServicio
        from clientes.models import Cliente
        
        lotificacion_id = request.query_params.get('lotificacion_id')
        
        configuraciones = ConfiguracionServicioLote.objects.filter(esta_activo=True)
        if lotificacion_id and lotificacion_id != 'all':
            configuraciones = configuraciones.filter(lote__manzana__lotificacion_id=lotificacion_id)
            
        lotes_ids = configuraciones.values_list('lote_id', flat=True)
        
        ventas = Venta.objects.filter(lote_id__in=lotes_ids).exclude(estado='CANCELADA')
        clientes_ids = ventas.values_list('cliente_id', flat=True).distinct()
        
        clientes_qs = Cliente.objects.filter(id__in=clientes_ids).order_by('nombres')
        
        paginator = StandardResultsSetPagination()
        page = paginator.paginate_queryset(clientes_qs, request)
        if page is not None:
            data = []
            for cliente in page:
                lotes_cliente = ventas.filter(cliente=cliente).values_list('lote_id', flat=True)
                
                pagos_atrasados = PagoServicio.objects.filter(
                    lote_id__in=lotes_cliente, 
                    estado__in=['Pendiente', 'Vencido']
                ).exists()
                
                estado_al_dia = not pagos_atrasados
                
                data.append({
                    'cliente_id': cliente.id,
                    'cliente_nombre': f"{cliente.nombres} {cliente.apellidos}".strip(),
                    'estado_al_dia': estado_al_dia
                })
                
            return paginator.get_paginated_response(data)
            
        return Response([])

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class LiquidacionComisionViewSet(viewsets.ModelViewSet):
    queryset = LiquidacionComision.objects.all()
    serializer_class = LiquidacionComisionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination

    def get_queryset(self):
        queryset = LiquidacionComision.objects.select_related('venta', 'vendedor', 'vendedor__empleado', 'venta__lote')
        
        # Filtros
        anio = self.request.query_params.get('anio')
        mes = self.request.query_params.get('mes')
        search = self.request.query_params.get('search')
        estado = self.request.query_params.get('estado')
        vendedor_id = self.request.query_params.get('vendedor')

        if anio and anio != 'all':
            queryset = queryset.filter(venta__fecha_creacion__year=anio)
        if mes and mes != 'all':
            queryset = queryset.filter(venta__fecha_creacion__month=mes)
        if estado and estado != 'all':
            queryset = queryset.filter(estado_pago=estado)
        if vendedor_id:
            queryset = queryset.filter(vendedor_id=vendedor_id)
        if search:
            queryset = queryset.filter(
                Q(vendedor__first_name__icontains=search) |
                Q(vendedor__last_name__icontains=search) |
                Q(vendedor__username__icontains=search) |
                Q(vendedor__empleado__nombre__icontains=search) |
                Q(vendedor__empleado__apellido__icontains=search)
            )

        return queryset.order_by('-fecha_creacion')

    @action(detail=True, methods=['post'])
    def pagar_ahora(self, request, pk=None):
        instance = self.get_object()
        if instance.estado_pago == 'PAGADO':
            return Response({'error': 'Esta liquidación ya ha sido pagada.'}, status=status.HTTP_400_BAD_REQUEST)
        
        forma = request.data.get('forma_pago')
        
        instance.estado_pago = 'PAGADO'
        instance.fecha_pago = timezone.now().date()
        instance.es_pago_inmediato = True
        instance.forma_pago = forma
        instance.save()
        
        return Response(self.get_serializer(instance).data)

    @action(detail=False, methods=['get'])
    def resumen_por_vendedor(self, request):
        anio = request.query_params.get('anio')
        mes = request.query_params.get('mes')
        
        if not anio or not mes or str(anio) == 'all' or str(mes) == 'all':
            return Response({'error': 'Debe especificar año y mes válidos.'}, status=status.HTTP_400_BAD_REQUEST)
            
        queryset = self.filter_queryset(self.get_queryset())
        
        from django.db.models import Sum, Count, Q, F
        
        resumen = queryset.values(
            'vendedor_id',
        ).annotate(
            vendedor_username=F('vendedor__username'),
            user_first_name=F('vendedor__first_name'),
            user_last_name=F('vendedor__last_name'),
            empleado_nombre=F('vendedor__empleado__nombre'),
            empleado_apellido=F('vendedor__empleado__apellido'),
            total_comisiones=Sum('monto_pagado'),
            monto_pendiente=Sum('monto_pagado', filter=Q(estado_pago='PENDIENTE')),
            monto_pagado_total=Sum('monto_pagado', filter=Q(estado_pago='PAGADO')),
            cantidad_ventas=Count('id'),
            cantidad_pendientes=Count('id', filter=Q(estado_pago='PENDIENTE'))
        ).order_by('-total_comisiones')
        
        resultados = []
        for item in resumen:
            if item.get('empleado_nombre'):
                nombre = f"{item['empleado_nombre']} {item.get('empleado_apellido') or ''}".strip()
            elif item.get('user_first_name'):
                nombre = f"{item['user_first_name']} {item.get('user_last_name') or ''}".strip()
            else:
                nombre = item.get('vendedor_username')
                
            resultados.append({
                'vendedor_id': item['vendedor_id'],
                'vendedor_nombre': nombre,
                'total_comisiones': float(item['total_comisiones'] or 0),
                'monto_pendiente': float(item['monto_pendiente'] or 0),
                'monto_pagado': float(item['monto_pagado_total'] or 0),
                'cantidad_ventas': item['cantidad_ventas'],
                'cantidad_pendientes': item['cantidad_pendientes']
            })
            
        return Response(resultados)

    @action(detail=False, methods=['post'])
    def pagar_multiples(self, request):
        ids = request.data.get('ids', [])
        forma = request.data.get('forma_pago', '')
        
        if not ids or not isinstance(ids, list):
            return Response({'error': 'Debe proporcionar una lista de IDs.'}, status=status.HTTP_400_BAD_REQUEST)
            
        liquidaciones = LiquidacionComision.objects.filter(id__in=ids, estado_pago='PENDIENTE')
        cantidad = liquidaciones.count()
        
        if cantidad == 0:
             return Response({'error': 'No se encontraron liquidaciones pendientes para los IDs proporcionados.'}, status=status.HTTP_400_BAD_REQUEST)
             
        liquidaciones.update(
            estado_pago='PAGADO',
            fecha_pago=timezone.now().date(),
            es_pago_inmediato=True,
            forma_pago=forma
        )
        
        return Response({'mensaje': f'Se registraron {cantidad} pagos exitosamente.'}, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset()).select_related(
            'venta__cliente', 'venta__lote__manzana__lotificacion', 'vendedor'
        )
        if not queryset.exists():
            return Response({'error': 'No hay registros para exportar en este periodo.'}, status=status.HTTP_400_BAD_REQUEST)
        
        wb = openpyxl.Workbook()
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        def aplicar_estilos(ws, title, headers, start_row=3):
            # Título principal
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
            ws.cell(row=1, column=1, value=title).font = Font(size=14, bold=True)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            
            # Headers
            ws.append([]) # Fila 2 vacía
            ws.append(headers) # Fila 3 headers
            
            for cell in ws[start_row]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
        def aplicar_bordes_y_anchos(ws, start_row=3):
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
            
            for col in ws.iter_cols(min_row=start_row, max_row=ws.max_row):
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = max_length + 2

        # ================================
        # HOJA 1: Ventas
        # ================================
        ws_ventas = wb.active
        ws_ventas.title = "Ventas"
        headers_v = ["Lotificacion", "Lote", "Vendedor", "Cliente", "Fecha", "Valor del Lote", "Valor de la venta", "Comision %", "Comision Q"]
        aplicar_estilos(ws_ventas, "Reporte de Ventas del Periodo", headers_v)
        
        ws_ventas.auto_filter.ref = f"A3:C{queryset.count()+3}"
        
        from empleados.models import Empleado
        
        resumen_dict = {}
        
        for liq in queryset:
            venta = liq.venta
            vendedor = liq.vendedor
            empleado = Empleado.objects.filter(usuario=vendedor).first()
            vendedor_nombre = f"{empleado.nombre} {empleado.apellido}".strip() if empleado else vendedor.username
            
            if vendedor_nombre not in resumen_dict:
                resumen_dict[vendedor_nombre] = {
                    'ventas': 0, 'total_vendido': 0, 'generada': 0, 'pagada': 0, 'pendiente': 0
                }
            resumen_dict[vendedor_nombre]['ventas'] += 1
            resumen_dict[vendedor_nombre]['total_vendido'] += float(venta.valor_lote)
            resumen_dict[vendedor_nombre]['generada'] += float(liq.monto_pagado)
            if liq.estado_pago == 'PAGADO':
                resumen_dict[vendedor_nombre]['pagada'] += float(liq.monto_pagado)
            else:
                resumen_dict[vendedor_nombre]['pendiente'] += float(liq.monto_pagado)

            porcentaje = empleado.porcentaje_comision if empleado else 0
            
            # Formatear Lote "A-1"
            lote_nombre = "N/A"
            if venta.lote:
                manz = venta.lote.manzana.nombre if venta.lote.manzana else ""
                lote_nombre = f"{manz}-{venta.lote.numero_lote}" if manz else venta.lote.numero_lote
            
            ws_ventas.append([
                venta.lote.manzana.lotificacion.nombre if venta.lote and venta.lote.manzana and venta.lote.manzana.lotificacion else "N/A",
                lote_nombre,
                vendedor_nombre,
                f"{venta.cliente.nombres} {venta.cliente.apellidos}",
                venta.fecha_creacion.strftime('%Y-%m-%d'),
                float(venta.lote.valor_total) if venta.lote else 0,
                float(venta.valor_lote),
                float(porcentaje),
                float(liq.monto_pagado)
            ])
            
        aplicar_bordes_y_anchos(ws_ventas)
        
        # ================================
        # HOJA 2: Pagos de Comisiones
        # ================================
        ws_pagos = wb.create_sheet(title="Pagos de Comisiones")
        headers_p = ["ID-Pago", "Folio Venta", "Vendedor", "Comision Q", "Fecha-Pago", "Estado"]
        aplicar_estilos(ws_pagos, "Reporte de Pagos Realizados", headers_p)
            
        ws_pagos.auto_filter.ref = f"C3:C{queryset.count()+3}"
        
        for liq in queryset:
            vendedor = liq.vendedor
            empleado = Empleado.objects.filter(usuario=vendedor).first()
            vendedor_nombre = f"{empleado.nombre} {empleado.apellido}".strip() if empleado else vendedor.username
            ws_pagos.append([
                liq.id,
                f"V-{liq.venta.id}",
                vendedor_nombre,
                float(liq.monto_pagado),
                liq.fecha_pago.strftime('%Y-%m-%d') if liq.fecha_pago else "",
                liq.estado_pago
            ])
        aplicar_bordes_y_anchos(ws_pagos)

        # ================================
        # HOJA 3: Resumen de Vendedores
        # ================================
        ws_resumen = wb.create_sheet(title="Resumen de Vendedores")
        headers_r = ["Vendedor", "Ventas", "Total Vendido", "Comision Generada", "Comision pagada", "Pendiente", "Firma"]
        aplicar_estilos(ws_resumen, "Resumen General por Vendedor", headers_r)
            
        for vendedor_nombre, data in resumen_dict.items():
            ws_resumen.append([
                vendedor_nombre,
                data['ventas'],
                data['total_vendido'],
                data['generada'],
                data['pagada'],
                data['pendiente'],
                ""
            ])
        aplicar_bordes_y_anchos(ws_resumen)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=planillas_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        return response

class CotizacionesPagination(PageNumberPagination):
    page_size = 10
    page_size_query_param = 'page_size'
    max_page_size = 100

class CotizacionViewSet(viewsets.ModelViewSet):
    queryset = Cotizacion.objects.all()
    serializer_class = CotizacionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CotizacionesPagination

    def get_queryset(self):
        user = self.request.user
        
        # Siempre restringir a las cotizaciones del usuario activo
        queryset = Cotizacion.objects.filter(vendedor=user).select_related('cliente', 'lote__manzana__lotificacion', 'vendedor')

        search = self.request.query_params.get('search')
        estado = self.request.query_params.get('estado')
        lotificacion = self.request.query_params.get('lotificacion')

        if estado:
            queryset = queryset.filter(estado=estado)
        
        if lotificacion:
            queryset = queryset.filter(lote__manzana__lotificacion_id=lotificacion)
        
        if not search and not estado and not lotificacion:
            # Si no hay búsqueda ni estado ni lotificación, dejamos el comportamiento base
            pass

        if search:
            queryset = queryset.filter(
                Q(cliente__nombres__icontains=search) |
                Q(cliente__apellidos__icontains=search) |
                Q(nombre_prospecto__icontains=search) |
                Q(lote__numero_lote__icontains=search)
            )

        return queryset.order_by('-fecha_creacion')

    def perform_create(self, serializer):
        serializer.save(vendedor=self.request.user)

    def perform_destroy(self, instance):
        if instance.estado == 'ACEPTADA':
            raise drf_serializers.ValidationError({'error': 'No se puede eliminar una cotización que ya fue aceptada y convertida en venta.'})
        instance.delete()

    @action(detail=True, methods=['post'])
    def convertir_a_venta(self, request, pk=None):
        cotizacion = self.get_object()

        # Validaciones de Estado
        if cotizacion.estado != 'PENDIENTE':
             return Response({'error': f'La cotización no se puede vender porque está en estado: {cotizacion.estado}.'}, status=status.HTTP_400_BAD_REQUEST)

        if cotizacion.es_vencida:
            cotizacion.estado = 'VENCIDA'
            cotizacion.save()
            return Response({'error': 'La cotización ha vencido.'}, status=status.HTTP_400_BAD_REQUEST)

        if not cotizacion.cliente:
            return Response({'error': 'La cotización no tiene un cliente asociado para crear la venta. Asigne un cliente primero.'}, status=status.HTTP_400_BAD_REQUEST)

        # Usar transacciones atómicas para cambiar estado
        from django.db import transaction
        with transaction.atomic():
            if cotizacion.lote.estado_disponibilidad != 'disponible':
                return Response({
                    'error': f'No se pudo concretar la cotización porque ya hay una venta registrada del lote cotizado (Lote {cotizacion.lote.numero_lote}).',
                    'code': 'LOTE_NO_DISPONIBLE'
                }, status=status.HTTP_400_BAD_REQUEST)

            try:
                nueva_venta = Venta(
                    cliente=cotizacion.cliente,
                    lote=cotizacion.lote,
                    valor_lote=cotizacion.valor_lote,
                    enganche=cotizacion.enganche,
                    descuento=cotizacion.descuento,
                    monto_financiar=cotizacion.monto_financiar,
                    tasa_interes_anual=cotizacion.tasa_interes_anual,
                    total_pagar_contado=cotizacion.total_pagar_contado,
                    tipo_pago=cotizacion.tipo_pago,
                    forma_pago=cotizacion.forma_pago,
                    acepta_instalacion=cotizacion.acepta_instalacion,
                    plazo_meses=cotizacion.plazo_meses,
                    vendedor=cotizacion.vendedor
                )
                nueva_venta.save()
                
                # Marcar cotización como ACEPTADA
                cotizacion.estado = 'ACEPTADA'
                cotizacion.save()
                
                return Response({'mensaje': 'Venta creada exitosamente', 'venta_id': nueva_venta.id}, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['post'])
    def rechazar(self, request, pk=None):
        cotizacion = self.get_object()
        if cotizacion.estado == 'ACEPTADA':
            return Response({'error': 'No se puede rechazar una cotización ya aceptada.'}, status=status.HTTP_400_BAD_REQUEST)
        
        cotizacion.estado = 'RECHAZADA'
        cotizacion.save()
        return Response({'mensaje': 'Cotización rechazada exitosamente.'})

    @action(detail=True, methods=['post'])
    def restaurar(self, request, pk=None):
        cotizacion = self.get_object()
        if cotizacion.estado != 'RECHAZADA':
            return Response({'error': 'Solo se pueden restaurar cotizaciones rechazadas.'}, status=status.HTTP_400_BAD_REQUEST)
        
        if cotizacion.es_vencida:
            cotizacion.estado = 'VENCIDA'
            cotizacion.save()
            return Response({'error': 'La cotización se restauró pero está VENCIDA por fecha.'}, status=status.HTTP_200_OK)
        
        cotizacion.estado = 'PENDIENTE'
        cotizacion.save()
        return Response({'mensaje': 'Cotización restaurada a PENDIENTE.'})

    @action(detail=True, methods=['get'])
    def exportar_excel(self, request, pk=None):
        cotizacion = self.get_object()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Cotizacion_{cotizacion.id}"

        # Estilos
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        border_side = Side(style='thin')
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        # Encabezado General
        ws.merge_cells('A1:E1')
        ws['A1'] = "COTIZACIÓN DE LOTE"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        # Información del Cliente y Vencimiento
        ws['A3'] = "Cliente/Prospecto:"
        ws['B3'] = f"{cotizacion.cliente.nombres} {cotizacion.cliente.apellidos}" if cotizacion.cliente else (cotizacion.nombre_prospecto or "Prospecto")
        ws['A4'] = "Teléfono:"
        ws['B4'] = (cotizacion.cliente.telefono if cotizacion.cliente else cotizacion.telefono_prospecto) or "N/A"
        ws['A5'] = "Fecha Vencimiento:"
        ws['B5'] = cotizacion.fecha_vencimiento.strftime('%d/%m/%Y')
        ws['A6'] = "Vendedor:"
        ws['B6'] = f"{cotizacion.vendedor.get_full_name() or cotizacion.vendedor.username}"

        # Información del Lote
        ws['D3'] = "Lote:"
        ws['E3'] = cotizacion.lote.numero_lote
        ws['D4'] = "Manzana:"
        ws['E4'] = cotizacion.lote.manzana.nombre
        ws['D5'] = "Área:"
        ws['E5'] = f"{cotizacion.lote.metros_cuadrados} m²"
        ws['D6'] = "Proyecto:"
        ws['E6'] = cotizacion.lote.manzana.lotificacion.nombre

        # Detalles Financieros
        ws['A8'] = "DETALLES FINANCIEROS"
        ws['A8'].font = Font(bold=True)
        
        ws['A9'] = "Valor del Lote:"
        ws['B9'] = float(cotizacion.valor_lote)
        ws['B9'].number_format = '\"Q\" #,##0.00'
        ws['A10'] = "Enganche:"
        ws['B10'] = float(cotizacion.enganche)
        ws['B10'].number_format = '\"Q\" #,##0.00'
        ws['A11'] = "Descuento:"
        ws['B11'] = float(cotizacion.descuento)
        ws['B11'].number_format = '\"Q\" #,##0.00'
        ws['A12'] = "Monto a Financiar:"
        ws['B12'] = float(cotizacion.monto_financiar or 0)
        ws['B12'].number_format = '\"Q\" #,##0.00'
        
        ws['D9'] = "Tipo de Pago:"
        ws['E9'] = cotizacion.tipo_pago
        ws['D10'] = "Forma de Pago:"
        ws['E10'] = cotizacion.forma_pago
        ws['D11'] = "Plazo (Meses):"
        ws['E11'] = cotizacion.plazo_meses
        ws['D12'] = "Tasa Interés Anual:"
        ws['E12'] = f"{cotizacion.tasa_interes_anual}%"

        # Cálculos de Interés Mensual y Cuota
        tasa_anual_decimal = float(cotizacion.tasa_interes_anual) / 100.0
        tasa_mensual_efectiva = round(tasa_anual_decimal / 12.0, 12)
        
        ws['D13'] = "Tasa Mensual EF.:"
        ws['E13'] = f"{round(tasa_mensual_efectiva * 100, 4)}%"

        # Amortización si es Financiado
        if cotizacion.tipo_pago == 'FINANCIADO' and cotizacion.plazo_meses > 0:
            row = 16
            ws.merge_cells(f'A{row}:F{row}')
            ws[f'A{row}'] = "TABLA DE AMORTIZACIÓN ESTIMADA"
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            row += 1
            headers = ["No.", "Fecha Est.", "Interés", "Capital", "Total Cuota", "Saldo"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
            
            row += 1
            saldo_insoluto = float(cotizacion.monto_financiar)
            fecha_vencimiento = timezone.now().date() + relativedelta(months=1)
            
            # Calcular cuota nivelada
            if tasa_mensual_efectiva > 0:
                numerador = saldo_insoluto * tasa_mensual_efectiva * math.pow(1 + tasa_mensual_efectiva, cotizacion.plazo_meses)
                denominador = math.pow(1 + tasa_mensual_efectiva, cotizacion.plazo_meses) - 1
                cuota_mensual = round(numerador / denominador, 2)
            else:
                cuota_mensual = round(saldo_insoluto / cotizacion.plazo_meses, 2)

            for i in range(1, cotizacion.plazo_meses + 1):
                interes_cuota = saldo_insoluto * tasa_mensual_efectiva
                capital_cuota = cuota_mensual - interes_cuota
                
                if i == cotizacion.plazo_meses:
                    capital_cuota = saldo_insoluto
                    cuota_mensual = capital_cuota + interes_cuota
                
                saldo_insoluto -= capital_cuota
                
                ws.cell(row=row, column=1, value=i).border = border
                ws.cell(row=row, column=2, value=fecha_vencimiento.strftime('%d/%m/%Y')).border = border
                
                c_interes = ws.cell(row=row, column=3, value=round(interes_cuota, 2))
                c_interes.number_format = '\"Q\" #,##0.00'
                c_interes.border = border
                
                c_capital = ws.cell(row=row, column=4, value=round(capital_cuota, 2))
                c_capital.number_format = '\"Q\" #,##0.00'
                c_capital.border = border
                
                c_total = ws.cell(row=row, column=5, value=round(cuota_mensual, 2))
                c_total.number_format = '\"Q\" #,##0.00'
                c_total.border = border
                
                c_saldo = ws.cell(row=row, column=6, value=max(0, round(saldo_insoluto, 2)))
                c_saldo.number_format = '\"Q\" #,##0.00'
                c_saldo.border = border
                
                fecha_vencimiento += relativedelta(months=1)
                row += 1

        # Ajustar columnas
        for col in ws.columns:
            max_length = 0
            # Usar el índice de columna y convertirlo a letra para evitar errores con MergedCells
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = max_length + 5

        user_filename = request.query_params.get('filename', f"cotizacion_{cotizacion.id}")
        if not user_filename.endswith('.xlsx'):
            user_filename += '.xlsx'

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{user_filename}"'
        wb.save(response)
        return response

